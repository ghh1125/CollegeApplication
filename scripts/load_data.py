"""Load existing data/raw CSV files into the local SQLite database."""

from __future__ import annotations

import csv
import json
import re
import shutil
import subprocess
import sys
import tempfile
from collections.abc import Iterator
from contextlib import contextmanager
from pathlib import Path
from typing import Any


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

RAW_DIR = PROJECT_ROOT / "data" / "raw"
HISTORICAL_CUTOFF_YEARS = (2025, 2024, 2023)
ADMISSION_PLAN_YEAR = 2025
ADMISSION_PROVINCE = "浙江"
DEFAULT_BATCH = "普通类"
DEFAULT_SUBJECT_REQUIREMENT = {"type": "NONE", "subjects": []}
DEFAULT_SUBJECT_REQUIREMENT_JSON = json.dumps(
    DEFAULT_SUBJECT_REQUIREMENT,
    ensure_ascii=False,
)

FIELD_ALIASES = {
    "school_code": ("学校代码", "院校代码", "学校代号", "院校代号", "院校编号"),
    "school_name": ("学校名称", "院校名称", "学校", "院校"),
    "major_code": (
        "专业代码",
        "专业代号",
        "专业(类)代码",
        "专业（类）代码",
        "专业编号",
    ),
    "major_name": (
        "专业名称",
        "专业(类)名称",
        "专业（类）名称",
        "专业",
        "专业类名称",
        "专业（类）名称",
    ),
    "plan_count": ("计划数", "招生计划数", "招生人数", "计划人数"),
    "subject_requirement": (
        "选考科目要求",
        "选考科目",
        "科目要求",
        "选考要求",
    ),
}
SUBJECT_NAMES = ("物理", "化学", "生物", "思想政治", "历史", "地理", "技术")
PDF_ROW_RE = re.compile(
    r"^(?P<province>[\u4e00-\u9fa5]{2,4})\s+"
    r"(?P<school>\S.+?)\s{2,}"
    r"(?P<major>\S.+?)\s{2,}"
    r".*?(?P<level>本科|高职\(专科\)|专科)\s+"
    r"(?P<requirement>(?:不提科目要求|[^\s]+\([^\n]+报考\)))\s*$"
)


def raw_csv_path(dataset: str, year: int, raw_dir: Path = RAW_DIR) -> Path:
    """Return the canonical raw CSV path for a dataset/year pair."""

    return raw_dir / f"{dataset}_{year}.csv"


def clean_text(value: Any) -> str:
    """Normalize scraped text cells."""

    text = str(value or "").strip()
    return re.sub(r"\s+", " ", text.replace("\u3000", " "))


def normalize_match_text(value: str) -> str:
    """Normalize school and major names for exact-ish matching."""

    text = clean_text(value)
    text = text.replace("（", "(").replace("）", ")")
    return re.sub(r"\s+", "", text)


def compact_major_name(value: str) -> str:
    """Drop parenthesized qualifiers for a conservative fallback key."""

    return re.sub(r"\([^)]*\)", "", normalize_match_text(value))


def subject_lookup_keys(school_name: str, major_name: str) -> list[tuple[str, str]]:
    """Return lookup keys for a school-major subject requirement pair."""

    school_key = normalize_match_text(school_name)
    major_key = normalize_match_text(major_name)
    compact_major_key = compact_major_name(major_name)
    keys = [(school_key, major_key)]
    if compact_major_key and compact_major_key != major_key:
        keys.append((school_key, compact_major_key))
    return keys


def read_csv_records(path: str | Path) -> list[dict[str, str]]:
    """Read CSV rows as dictionaries."""

    with Path(path).open(encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def pick(row: dict[str, Any], field: str, default: str = "") -> str:
    """Return the first matching field value from a source row."""

    for key in FIELD_ALIASES[field]:
        value = row.get(key)
        if value is not None and str(value).strip():
            return clean_text(value)
    return default


def parse_int(value: Any) -> int | None:
    """Parse integer-like scraped values such as plan counts."""

    text = clean_text(value)
    if not text or text in {"-", "--", "—"}:
        return None
    match = re.search(r"-?\d+", text.replace(",", ""))
    return int(match.group(0)) if match else None


def subject_requirement_json_from_text(text: str | None) -> str:
    """Convert a subject requirement string into the JSON shape used by SQLite."""

    requirement = clean_text(text)
    if not requirement or "不提科目要求" in requirement or "不限" in requirement:
        return DEFAULT_SUBJECT_REQUIREMENT_JSON

    subjects = [subject for subject in SUBJECT_NAMES if subject in requirement]
    if "均须" in requirement or "必须同时" in requirement:
        requirement_type = "ALL"
    elif "其中" in requirement or "任意" in requirement:
        requirement_type = "ANY"
    elif len(subjects) == 1:
        requirement_type = "ALL"
    else:
        requirement_type = "TEXT"
    return json.dumps(
        {"type": requirement_type, "subjects": subjects},
        ensure_ascii=False,
    )


def spreadsheet_records_from_xlsx(path: Path) -> list[dict[str, str]]:
    """Read subject requirement records from .xlsx or legacy .xls bytes."""

    content = path.read_bytes()
    if content.startswith(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"):
        from scripts.fetch_data import spreadsheet_bytes_to_records

        return spreadsheet_bytes_to_records(content)

    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = [
        [clean_text(cell) for cell in row]
        for row in sheet.iter_rows(values_only=True)
        if any(clean_text(cell) for cell in row)
    ]
    header_index = 0
    for index, row in enumerate(rows[:30]):
        if any("院校" in cell or "学校" in cell for cell in row) and any(
            "专业" in cell for cell in row
        ):
            header_index = index
            break
    headers = rows[header_index]
    records: list[dict[str, str]] = []
    for row in rows[header_index + 1 :]:
        padded = row + [""] * (len(headers) - len(row))
        records.append(
            {
                headers[index] or f"column_{index + 1}": padded[index]
                for index in range(len(headers))
            }
        )
    return records


def subject_requirement_records_from_pdf(path: Path) -> list[dict[str, str]]:
    """Extract best-effort subject requirement records from the official PDF."""

    pdftotext = shutil.which("pdftotext")
    if pdftotext is None:
        return []
    with tempfile.TemporaryDirectory() as tmpdir:
        text_path = Path(tmpdir) / "subject_requirement.txt"
        result = subprocess.run(
            [pdftotext, "-layout", str(path), str(text_path)],
            check=False,
            capture_output=True,
            text=True,
        )
        if result.returncode != 0 or not text_path.exists():
            return []
        lines = text_path.read_text(encoding="utf-8", errors="ignore").splitlines()

    records: list[dict[str, str]] = []
    for line in lines:
        if (
            "浙江省教育考试院" in line
            or "第 " in line
            or "省份" in line
            or not line.strip()
        ):
            continue
        match = PDF_ROW_RE.match(line.rstrip())
        if not match:
            continue
        records.append(
            {
                "院校名称": clean_text(match.group("school")),
                "专业(类)名称": clean_text(match.group("major")),
                "选考科目要求": clean_text(match.group("requirement")),
            }
        )
    return records


def load_subject_requirement_index(raw_dir: Path = RAW_DIR) -> dict[tuple[str, str], str]:
    """Load subject requirements indexed by normalized school and major name."""

    xlsx_path = raw_dir / "subject_requirement.xlsx"
    pdf_path = raw_dir / "subject_requirement.pdf"
    if xlsx_path.exists():
        records = spreadsheet_records_from_xlsx(xlsx_path)
    elif pdf_path.exists():
        records = subject_requirement_records_from_pdf(pdf_path)
    else:
        records = []

    index: dict[tuple[str, str], str] = {}
    for record in records:
        school_name = pick(record, "school_name")
        major_name = pick(record, "major_name")
        requirement = pick(record, "subject_requirement")
        if not (school_name and major_name and requirement):
            continue
        for key in subject_lookup_keys(school_name, major_name):
            index.setdefault(key, requirement)
    return index


def find_subject_requirement(
    subject_index: dict[tuple[str, str], str],
    school_name: str,
    major_name: str,
) -> str | None:
    """Find a subject requirement for a school-major pair."""

    for key in subject_lookup_keys(school_name, major_name):
        requirement = subject_index.get(key)
        if requirement:
            return requirement
    return None


@contextmanager
def connection_scope(conn: Any | None = None) -> Iterator[Any]:
    """Use an existing connection in tests or open a managed DB connection."""

    if conn is not None:
        yield conn
        return

    from db import get_conn

    with get_conn() as managed_conn:
        yield managed_conn


def insert_admission_plan_from_cutoff(
    cutoff_path: str | Path,
    year: int,
    subject_index: dict[tuple[str, str], str],
    conn: Any,
) -> dict[str, int]:
    """Use the selected year's cutoff CSV as the current admission-plan base."""

    from db import get_cursor

    sql = """
        INSERT INTO admission_plan (
            year,
            province,
            batch,
            recruit_type,
            school_code,
            school_name,
            major_code,
            major_name,
            plan_count,
            subject_requirement,
            subject_requirement_text,
            subject_requirement_json,
            school_location,
            tuition
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT (year, province, batch, school_code, major_code)
        DO UPDATE SET
            school_name = EXCLUDED.school_name,
            major_name = EXCLUDED.major_name,
            plan_count = EXCLUDED.plan_count,
            subject_requirement = EXCLUDED.subject_requirement,
            subject_requirement_text = EXCLUDED.subject_requirement_text,
            subject_requirement_json = EXCLUDED.subject_requirement_json,
            school_location = EXCLUDED.school_location,
            tuition = EXCLUDED.tuition
    """

    rows = read_csv_records(cutoff_path)
    inserted = 0
    matched = 0
    unmatched = 0
    with get_cursor(conn) as cursor:
        for row in rows:
            school_code = pick(row, "school_code")
            school_name = pick(row, "school_name")
            major_code = pick(row, "major_code")
            major_name = pick(row, "major_name")
            if not (school_code and school_name and major_code and major_name):
                continue
            requirement_text = find_subject_requirement(
                subject_index,
                school_name,
                major_name,
            )
            if requirement_text:
                matched += 1
            else:
                unmatched += 1
            cursor.execute(
                sql,
                (
                    year,
                    ADMISSION_PROVINCE,
                    DEFAULT_BATCH,
                    "MAJOR",
                    school_code,
                    school_name,
                    major_code,
                    major_name,
                    parse_int(pick(row, "plan_count")),
                    requirement_text,
                    requirement_text,
                    subject_requirement_json_from_text(requirement_text),
                    None,
                    None,
                ),
            )
            inserted += 1
    return {
        "admission_plan_rows": inserted,
        "subject_req_matched": matched,
        "subject_req_unmatched": unmatched,
    }


def run_ingestion(
    raw_dir: Path = RAW_DIR,
    ingest_module: Any | None = None,
    conn: Any | None = None,
) -> dict[str, Any]:
    """Load cutoff CSVs and derive the 2025 admission plan into SQLite."""

    if ingest_module is None:
        from app.pipeline import ingest as ingest_module

    plan_base_path = raw_csv_path("historical_cutoff", ADMISSION_PLAN_YEAR, raw_dir)
    if not plan_base_path.exists():
        raise FileNotFoundError(f"missing 2025 cutoff CSV: {plan_base_path}")

    subject_index = load_subject_requirement_index(raw_dir)
    if not subject_index:
        print(
            "警告：未找到可解析的 subject_requirement.xlsx，"
            "选考科目默认写入 {\"type\": \"NONE\", \"subjects\": []}"
        )

    stats: dict[str, Any] = {"historical_cutoff_rows": {}}
    with connection_scope(conn) as active_conn:
        stats.update(
            insert_admission_plan_from_cutoff(
                plan_base_path,
                ADMISSION_PLAN_YEAR,
                subject_index,
                active_conn,
            )
        )
        for year in HISTORICAL_CUTOFF_YEARS:
            cutoff_path = raw_csv_path("historical_cutoff", year, raw_dir)
            if not cutoff_path.exists():
                raise FileNotFoundError(f"missing historical cutoff CSV: {cutoff_path}")
            stats["historical_cutoff_rows"][year] = ingest_module.ingest_historical_cutoff(
                str(cutoff_path),
                year=year,
                conn=active_conn,
            )
        ingest_module.build_program_mapping(year=ADMISSION_PLAN_YEAR, conn=active_conn)
    return stats


def load_validation_summary(conn: Any) -> str:
    """Run verification SQL queries and format their output."""

    from db import get_cursor

    with get_cursor(conn) as cursor:
        cursor.execute("SELECT year, COUNT(*) FROM admission_plan GROUP BY year ORDER BY year")
        plan_rows = cursor.fetchall()
    with get_cursor(conn) as cursor:
        cursor.execute(
            "SELECT year, COUNT(*) FROM historical_cutoff GROUP BY year ORDER BY year"
        )
        cutoff_rows = cursor.fetchall()
    with get_cursor(conn) as cursor:
        cursor.execute(
            "SELECT "
            "SUM(CASE WHEN subject_requirement_text IS NOT NULL "
            "AND subject_requirement_text != '' THEN 1 ELSE 0 END), "
            "SUM(CASE WHEN subject_requirement_text IS NULL "
            "OR subject_requirement_text = '' THEN 1 ELSE 0 END) "
            "FROM admission_plan WHERE year = 2025"
        )
        subject_row = cursor.fetchone() or (0, 0)

    plan_text = " / ".join(f"{year}年 {count} 条" for year, count in plan_rows) or "无数据"
    cutoff_text = (
        " / ".join(f"{year}年 {count} 条" for year, count in cutoff_rows) or "无数据"
    )
    matched_count = subject_row[0] or 0
    unmatched_count = subject_row[1] or 0
    return "\n".join(
        (
            "=== 数据入库验证 ===",
            f"admission_plan:    {plan_text}",
            f"historical_cutoff: {cutoff_text}",
            f"subject_req matched: {matched_count} 条 / 未匹配: {unmatched_count} 条",
        )
    )


def print_validation_summary() -> None:
    """Print database verification counts."""

    from db import get_conn

    with get_conn() as conn:
        print(load_validation_summary(conn))


def main() -> None:
    """Create tables, ingest raw CSVs, and print verification counts."""

    from scripts.init_db import initialize_database

    initialize_database()
    run_ingestion(RAW_DIR)
    print_validation_summary()


if __name__ == "__main__":
    main()
