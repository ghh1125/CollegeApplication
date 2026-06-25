"""Regression tests for page form collection helpers."""

from __future__ import annotations

import importlib
import sys
from types import ModuleType
from types import SimpleNamespace
import unittest


class UICollectFormTests(unittest.TestCase):
    """Each province page should expose the form context used by AI fill."""

    def setUp(self) -> None:
        self._old_streamlit = sys.modules.get("streamlit")
        fake = ModuleType("streamlit")
        fake.session_state = {}
        sys.modules["streamlit"] = fake

    def tearDown(self) -> None:
        for name in ("ui.zhejiang_page", "ui.questionnaire"):
            sys.modules.pop(name, None)
        if self._old_streamlit is None:
            sys.modules.pop("streamlit", None)
        else:
            sys.modules["streamlit"] = self._old_streamlit

    def _import_page(self, name: str):  # noqa: ANN202
        return importlib.import_module(f"ui.{name}_page")

    def _with_state(self, module, state: dict) -> dict:  # noqa: ANN001
        original_st = module.st
        module.st = SimpleNamespace(session_state=state)
        try:
            return module._collect_form()
        finally:
            module.st = original_st

    def test_zhejiang_collect_form_has_defaults(self) -> None:
        zhejiang_page = self._import_page("zhejiang")

        form = self._with_state(zhejiang_page, {})

        self.assertEqual(form["rank"], 8000)
        self.assertEqual(form["selected_subjects"], [])
        self.assertEqual(form["main_priority"], "请选择…")
        self.assertEqual(form["preferred_majors"], [])

    def test_zhejiang_school_name_link_attaches_admission_url(self) -> None:
        zhejiang_page = self._import_page("zhejiang")

        linked = zhejiang_page._school_name_link("https://bkzs.pku.edu.cn/", "北京大学")

        self.assertTrue(linked.startswith("https://bkzs.pku.edu.cn/"))
        self.assertIn("__school_name__=北京大学", linked)

    def test_zhejiang_excel_export_includes_risk_pools_and_new_major_sheet(self) -> None:
        import io

        import pandas as pd
        from openpyxl import load_workbook

        zhejiang_page = self._import_page("zhejiang")
        data = zhejiang_page._to_excel(
            pd.DataFrame([{"序号": 1, "专业名称": "计算机科学与技术"}]),
            [{"序号": 1, "冲稳保": "冲"}],
            [{"序号": 1, "冲稳保": "稳"}],
            [{"序号": 1, "冲稳保": "保"}],
            [{"序号": 1, "专业名称": "未来技术"}],
            ["序号", "专业名称"],
            ["序号", "专业名称"],
        )

        workbook = load_workbook(io.BytesIO(data), read_only=True)

        self.assertEqual(
            workbook.sheetnames,
            ["推荐80志愿", "冲候选池", "稳候选池", "保候选池", "2026新专业"],
        )


if __name__ == "__main__":
    unittest.main()
